import re
import os
from datetime import datetime, timezone
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ---- CONFIG ----
INPUT_XLSX = "data/input/Price Comparison - 14Jan'26.xlsx"   # your current file in repo
OUTPUT_XLSX = "data/output/price_output.xlsx"                # ONLY output file (fixed name)
SHEET_NAME = "Combined"                                      # change if your sheet differs

LINK_COL = 4  # Column D
# ----------------

PRICE_SELECTORS = [
    "#corePriceDisplay_desktop_feature_div span.a-offscreen",
    "#corePriceDisplay_mobile_feature_div span.a-offscreen",
    "span.a-price span.a-offscreen",
    "#priceblock_ourprice",
    "#priceblock_dealprice",
]

BLOCK_MARKERS = [
    "Robot Check",
    "Enter the characters you see below",
    "Sorry, we just need to make sure you're not a robot",
]

def get_link_target(cell):
    # Prefer actual hyperlink target (Excel hyperlink), fallback to visible text if it's a URL
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    val = (str(cell.value).strip() if cell.value is not None else "")
    return val if val.startswith("http") else ""

def parse_price(text: str) -> str:
    if not text:
        return ""
    t = text.replace("₹", "").replace(",", "").strip()
    t = re.sub(r"[^0-9.]", "", t)
    return t

def first_text(page, selectors):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0:
                txt = loc.inner_text(timeout=1500).strip()
                if txt:
                    return txt
        except Exception:
            pass
    return ""

def main():
    wb = load_workbook(INPUT_XLSX)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # Insert ONE new column right after D (i.e., new column E)
    price_col = LINK_COL + 1
    ws.insert_cols(price_col)
    ws.cell(row=1, column=price_col).value = "Amazon Price (INR)"

    # Optional: add timestamp in A1 so you can see file is fresh
    ws.cell(row=1, column=1).value = f"Last updated (UTC): {datetime.now(timezone.utc).isoformat()}"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            locale="en-IN",
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/122.0.0.0 Safari/537.36"),
        )
        page = context.new_page()

        for r in range(2, ws.max_row + 1):
            link_cell = ws.cell(row=r, column=LINK_COL)
            url = get_link_target(link_cell)

            if not url:
                ws.cell(row=r, column=price_col).value = ""
                continue

            # Only scrape amazon.in links; skip others
            if "amazon.in" not in url:
                ws.cell(row=r, column=price_col).value = ""
                continue

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=45000)
                page.wait_for_timeout(1500)

                title = (page.title() or "")
                html = page.content()

                if any(m.lower() in title.lower() for m in BLOCK_MARKERS) or any(m.lower() in html.lower() for m in BLOCK_MARKERS):
                    price = ""  # blocked
                else:
                    price_text = first_text(page, PRICE_SELECTORS)
                    price = parse_price(price_text)

                ws.cell(row=r, column=price_col).value = price

            except PWTimeout:
                ws.cell(row=r, column=price_col).value = ""
            except Exception:
                ws.cell(row=r, column=price_col).value = ""

        context.close()
        browser.close()

    wb.save(OUTPUT_XLSX)
    print(f"Saved: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
